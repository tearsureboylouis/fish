import os.path
import openpyxl


def convert_to_center_coordinates(x, y):
    """计算每条鱼在现实中的以声呐为原点的坐标中中的距离和位置，单位为cm"""
    center_x = (x - 0.5) * (600 / 604) * 804
    center_y = (y - 0.5) * 600
    return center_x, center_y


def get_fish_coordinates(base_path, current_ping):
    coordinates = []
    filename = os.path.join(base_path, 'metadata', f'{current_ping}.txt')
    try:
        with open(filename, 'r', encoding='utf-8') as file:
            lines = file.readlines()
            for line in lines:
                numbers = line.split()
                if len(numbers) >= 3:
                    x_left_top = float(numbers[1])
                    y_left_top = float(numbers[2])
                    x, y = convert_to_center_coordinates(x_left_top, y_left_top)
                    coordinates.append((x, y))
            return coordinates
    except FileNotFoundError:
        return []


def out_put_result(base_path, total_frame, fps, duration, depth):
    """Excel表格设置"""
    output_coordinates_file = os.path.join(base_path, 'results/coordinates.xlsx')
    output_depth_fish_file = os.path.join(base_path, 'results/count_deep_second.xlsx')
    output_3d_coordinates_file = os.path.join(base_path, 'results/deep_frame_coordinates.xlsx')

    workbook_coordinates = openpyxl.Workbook()
    sheet_coordinates = workbook_coordinates.active
    sheet_coordinates.title = 'Coordinates'
    sheet_coordinates.append(['x', 'y'])

    workbook_depth_fish = openpyxl.Workbook()
    sheet_depth_fish = workbook_depth_fish.active
    sheet_depth_fish.title = 'Fish Count Deep'
    sheet_depth_fish.append(['Time (s)', 'Fish Count', 'Depth (cm)'])

    workbook_3d_coordinates = openpyxl.Workbook()
    sheet_3d_coordinates = workbook_3d_coordinates.active
    sheet_3d_coordinates.title = '3D Coordinates'
    sheet_3d_coordinates.append(['Time (ping)', 'Fish Count', 'Depth (cm)', 'Coordinates'])

    frame_rate = 0
    row_video_duration = duration
    water_depth = depth
    total_seconds = (total_frame / fps)

    if row_video_duration != 0:
        frame_rate = round(fps * (total_seconds / row_video_duration))
    time = ((row_video_duration / total_seconds) * (1 / fps))

    fish_counts = []

    n = 1

    for current_ping in range(total_frame):
        coordinates = get_fish_coordinates(base_path, current_ping)
        for coord in coordinates:
            sheet_coordinates.append(list(coord))
        count = len(coordinates)

        fish_counts.append(count)

        """Fish Count Deep(sheet_depth_fish)，添加不同深度的鱼"""
        if frame_rate != 0 and current_ping % frame_rate == 0:
            depth = (water_depth / row_video_duration) * (n)
            sheet_depth_fish.append([n, count, depth])
            n += 1

        """读取 frame_rate_H 个 txt 文件，将坐标汇总到 coordinates_list"""
        coordinates_list = []

        count = fish_counts[current_ping]
        coordinates_list.extend(coordinates)

        # 过滤掉相同的坐标
        unique_coordinates = list(set(coordinates_list))
        depth = (water_depth / row_video_duration) * (time) * (current_ping)
        # 将这些坐标作为一类，导出到新的 3D_coordinates.xlsx 文件
        sheet_3d_coordinates.append([f'{current_ping}', count, depth, str(unique_coordinates)])

    workbook_coordinates.save(output_coordinates_file)
    workbook_depth_fish.save(output_depth_fish_file)
    workbook_3d_coordinates.save(output_3d_coordinates_file)
